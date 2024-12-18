import openpyxl

# 엑셀 파일 경로를 설정합니다.
file_path = "11월 사용량 정리.xlsx"

# 엑셀 파일 불러오기
wb = openpyxl.load_workbook(file_path)

# 1번 시트 기준
main_sheet = wb["전체"]

sheet_names = wb.sheetnames

target_sheets = [str(i) for i in range(2, 15)]

# "전체" 시트의 H~T열 설정
output_columns = [chr(col) for col in range(ord('H'), ord('V') + 1)]

# 로그 출력
print("### 데이터 찾기 및 입력 시작 ###")

# '전체' 시트의 C열 (2번째 행부터 231번째 행까지) 값 기준으로 반복
for row in range(2, 232):  # 2번째 행부터 231번째 행까지
    search_value = main_sheet[f'C{row}'].value  # 찾을 문자열 (C열)
    if search_value is None:
        continue  # C열이 비어있으면 건너뜀
    search_value = str(search_value).strip().lower()  # 문자열 정제 (공백 제거 및 소문자화)

    print(f"\n[LOG] '{search_value}' 값을 찾습니다. (행: {row})")
    
    column_index = 0  # H~T 열을 순서대로 사용하기 위한 인덱스

    # 시트 순회 (2번부터 14번 시트)
    for sheet_name in sheet_names:
        if column_index >= len(output_columns):
            print(f"    [WARNING] 열 개수(H~T)를 초과했습니다. 추가 데이터는 무시됩니다.")
            break  # 열 개수를 초과하면 중단
        
        sheet = wb[sheet_name]  # 시트 불러오기
        found_value = 0  # 초기값은 0

        # 해당 시트의 A열을 검사
        for sheet_row in range(2, sheet.max_row + 1):
            a_value = sheet[f'A{sheet_row}'].value
            if a_value is None:
                continue
            a_value = str(a_value).strip().lower()  # 비교 대상 문자열 정제

            if a_value == search_value:  # 값이 일치하면
                found_value = sheet[f'C{sheet_row}'].value  # C열 값 가져오기
                if found_value is None:  # C열 값이 비어있으면 0으로 처리
                    found_value = 0
                print(f"    [FOUND] 시트 '{sheet_name}' (행: {sheet_row})에서 '{found_value}' 값을 찾음.")
                break  # 일치하는 값을 찾으면 중단
        
        # 결과를 '전체' 시트의 H~T 열에 순서대로 입력
        main_sheet[f'{output_columns[column_index]}{row}'] = found_value
        print(f"    [INPUT] '{output_columns[column_index]}{row}'에 '{found_value}' 입력.")
        column_index += 1  # 다음 열로 이동

# 수정된 엑셀 파일 저장
wb.save("합산.xlsx")
print("합산 완료 및 저장되었습니다.")